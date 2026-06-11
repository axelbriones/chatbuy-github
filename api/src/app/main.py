from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
import boto3
from fpdf import FPDF, HTML2FPDF
from jinja2 import Environment
from urllib.parse import quote
from boto3.dynamodb.conditions import Key, Attr
import os
import time
from datetime import datetime
import openpyxl
from openpyxl.styles import Font 
from weasyprint import HTML, CSS

AWS_ACCESS_KEY_ID = "AKIAWD5OETVVWTKP3SW6"
AWS_SECRET_ACCESS_KEY = "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"
s3_bucket = "qullqua-uploads"


app = FastAPI()
origins = ["*"]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

dynamodb = boto3.resource('dynamodb', region_name='us-east-1', 
                          aws_access_key_id=AWS_ACCESS_KEY_ID,
                          aws_secret_access_key=AWS_SECRET_ACCESS_KEY)
s3 = boto3.client('s3', region_name='us-east-1', 
                          aws_access_key_id=AWS_ACCESS_KEY_ID,
                          aws_secret_access_key=AWS_SECRET_ACCESS_KEY)


table = dynamodb.Table('collector-atwork-pro')
table_country = dynamodb.Table('collector-country-pro')
table_tecnique = dynamodb.Table('collector-technique-pro')
table_category = dynamodb.Table('collector-category-pro')
table_art = dynamodb.Table('collector-artist-pro')
table_coll = dynamodb.Table('collector-collection-pro')
table_user = dynamodb.Table('collector-users-pro')





def build_simple_xls(items, tipo,name="anonimo"):
    timestamp = int(time.time())
    wb = openpyxl.Workbook()
    ws = wb.active
    data = []
    for item in items:
        if tipo == "1":
            dato = {"Titulo":  item.get('atname','') if item.get('atname') is not None else "",
                    "Descripcion":  item.get('description','') if item.get('description') is not None else "", 
                    "Categoria":  item.get('category_name','') if item.get('category_name') is not None else "", 
                    "Tecnica": item.get('tecnique_name','') if item.get('tecnique_name') is not None else "",
                    "Author": item.get('author','') if item.get('author') is not None else "" }
        else:
             
             dato = {
                    "id" : item.get('id','') if item.get('id') is not None else "",
                    'Coleccion':  item.get('collection_name','') if item.get('collection_name') is not None else "",
                    'Categoria':  item.get('category_name','') if item.get('category_name') is not None else "",
                    'Tecnica': item.get('tecnique_name','') if item.get('tecnique_name') is not None else "",
                    'Autor': item.get('author','') if item.get('author') is not None else "",
                    'Estatus':  item.get('artwork_status','') if item.get('artwork_status') is not None else "",
                    'FotoMasInfo':  item.get('photoMasInfo','') if item.get('photoMasInfo') is not None else "",
                    'FotoCertificado':  item.get('photoCertificado','') if item.get('photoCertificado') is not None else "",
                    'FotoBiografia':  item.get('photoBiografia','') if item.get('photoBiografia') is not None else "",
                    'Foto': item.get('photo','') if item.get('photo') is not None else "",
                    'Conservacion':  item.get('conservation','') if item.get('conservation') is not None else "",
                    'Nota':  item.get('note','') if item.get('note') is not None else "",
                    'Warrant':  item.get('warrant','') if item.get('warrant') is not None else "",
                    'PrecioMarket':  item.get('price_market','') if item.get('price_market') is not None else "",
                    'PrecioAdquisicion': item.get('price_adquisition','') if item.get('price_adquisition') is not None else "",
                    'FechaAdquisicion':  item.get('date_adquisition','') if item.get('date_adquisition') is not None else "",
                    'Adquisicion':  item.get('adquisition','') if item.get('adquisition') is not None else "",
                    'Lugar':   item.get('artwork_location','') if item.get('artwork_location') is not None else "",
                    'TamaDes': item.get('sizes_des','') if item.get('sizes_des') is not None else "",
                    'Pais':  item.get('country_name','') if item.get('country_name') is not None else "",
                    'AñoAt':  item.get('year_at','') if item.get('year_at') is not None else "",
                    'Descripcion':  item.get('description','') if item.get('description') is not None else "", 
                    'DescripcionShort':  item.get('DescripcionShort','') if item.get('DescripcionShort') is not None else "",
                    'Código': item.get('register_code','') if item.get('register_code') is not None else "",
                    'Nombre': item.get('atname','') if item.get('atname') is not None else "",
                    'Certificado': item.get('certificate_status','') if item.get('certificate_status') is not None else "",
                    'Activo': item.get('active','') if item.get('active') is not None else "",
                    'Moneda':  item.get('currency_symbol','') if item.get('currency_symbol') is not None else "",
                    'Titulo': item.get('atname','') if item.get('atname') is not None else "",
                    'Activa':  item.get('active','') if item.get('active') is not None else ""
                }

        data.append(dato)


    bold_font = Font(bold=True)
    column_headers = data[0].keys()

    for col_num, header in enumerate(column_headers):
        ws.cell(row=1, column=col_num + 1).value = header
        ws.cell(row=1, column=col_num + 1).font = bold_font


    for row_num, row_data in enumerate(data):
        
        current_row = row_num + 2
        for col_num, (key, value) in enumerate(row_data.items()):
    
            current_col = col_num + 1  
            ws.cell(row=current_row, column=current_col).value = value


    if tipo == "1":
        filename = f"Resumen-simple-{name}-{timestamp}.xls"
    else:
         filename = f"Resumen-total-{name}-{timestamp}.xls"

    wb.save(filename)
    s3.upload_file(filename, s3_bucket, filename, ExtraArgs={'ACL': 'public-read'})
    s3_url = s3.generate_presigned_url(
                        ClientMethod="get_object",
                        Params={"Bucket": s3_bucket, "Key": filename}
                    )
    os.remove(filename)
    
    return s3_url.split("?")[0]

def buil_pdf_inventory(data, tipo,name="anonimo"):
    template_list = Environment().from_string(
        """
        {% block head %}
            <sup><i>{{ now }}</i></sup>
            <div>                        
                <p><strong>Inventario {{ tipo }}</strong></p>
            </div>
        {% endblock %}
        <body>
        {% for title, items in data.items() %}
            · <title><b>{{ title | escape }}</b></title>
            <ul>
            {% for item in items %}
                <li> - {{ item.name }} ({{ item.count }})</li>
            {% endfor %}
            </ul>
        {% endfor %}
        </body>
        """
    )
    timestamp = int(time.time())
    filename = f"Inventario-Obras-{tipo}-{name}-{timestamp}.pdf"
    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("Montserrat", fname="/usr/src/app/app/Montserrat-Regular.ttf")
    pdf.add_font("Montserrat", fname="/usr/src/app/app/Montserrat-Bold.ttf", style="B")
    pdf.add_font("Montserrat", fname="/usr/src/app/app/Montserrat-Italic.ttf", style="i")
    pdf.set_font("Montserrat", "i", size=8)
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    
    pdf.write_html(template_list.render(data=data,now=now,tipo=f"Inventario-Obras-{tipo}"))
    pdf.output(filename)

    s3.upload_file(filename, s3_bucket, filename, ExtraArgs={"ACL": "public-read"})
    s3_url = s3.generate_presigned_url(
        ClientMethod="get_object", Params={"Bucket": s3_bucket, "Key": filename}
    )
    os.remove(filename)
    return s3_url.split("?")[0]


def buil_pdf(data, name="anonimo"):
    # Plantilla HTML
    template_list = Environment().from_string("""
    <html>
    <head>
        <style>
            @page {
                size: A4 landscape; /* Establecer tamaño de página A4 en formato horizontal */
                margin: 6mm; /* Margen opcional */
            }
            @font-face {
                font-family: 'Montserrat';
                src: url('Montserrat-Regular.ttf') format('truetype'); 
            }
            body {
                font-family: 'Montserrat', sans-serif;
                font-size: 14px; 
            }
            table { width: 100%; border-collapse: collapse; }
            th {
                border-bottom: 1px solid black;
                padding: 8px; 
                text-align: center; 
                vertical-align: center; 
                width: 100px;
            }
            td { 
                border-bottom: 1px solid black;
                padding: 8px; 
                text-align: center; 
                vertical-align: center; 
                overflow-wrap: break-word;
                text-overflow: elipsis;
                width: 100px;
                height: 90px;
            }
            img { height: 80px; max-width: 80px; /* Limitar tamaño de la imagen */ }
        </style>
    </head>
    <body>
        <sup><i>{{ now }}</i></sup>
        <h1>Resumen de Obras</h1>
        <table>
            <thead>
                <tr>
                    <th>Imagen</th>
                    <th>Artista</th>
                    <th>Medio</th>
                    <th>Nombre</th>
                    <th>Descripción</th>
                    <th>Año</th>
                    <th>Pais</th>
                    <th>Medidas</th>
                    <th>Ubicación obra</th>
                </tr>
            </thead>
            <tbody>
                {% for obra in obras %}
                    <tr>
                        <td><img src="{{ obra.imagen | escape }}" alt="imagen"></td>
                        <td>{{ obra.artista }}</td>
                        <td>{{ obra.medio }}</td>
                        <td>{{ obra.nombre }}</td>
                        <td>{{ obra.descripcion }}</td>
                        <td style="max-width: 70px;">{{ obra.año }}</td>
                        <td>{{ obra.pais }}</td>
                        <td style="max-width: 70px;">{{ obra.medidas }}</td>
                        <td>{{ obra.ubicacion_obra }}</td>
                    </tr>
                {% endfor %}                       
            </tbody>
        </table>
    </body>
    </html>
    """)

    obras = []
    for item in data:
        image_url = item.get('photo_thumb', 'https://qullqua-uploads.s3.amazonaws.com/blanco2.jpeg')
        encoded_image_url = quote(image_url, safe='/:')
        description = item.get('description', '')
        description = description[:50] if description else ''
        medidas = item.get('sizes_des', '')
        medidas = medidas[:40] if medidas else ''
        
        item2 = {
            "descripcion": description,
            "medidas": medidas,
            "imagen": encoded_image_url,
            "medio": item.get('tecnique_name', ''),
            "artista": item.get('author', ''),
            "nombre": item.get('atname', ''),
            "año": item.get('year_at', ''),
            "pais": item.get('country_name', ''),
            "ubicacion_obra": item.get('artwork_location', '')
        }
        obras.append(item2)

    # Ordenar obras por artista (de A a Z)
    obras.sort(key=lambda x: (x['artista'] or '').lower())

    # Generar HTML con Jinja2
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    rendered_html = template_list.render(obras=obras, now=now)

    # Generar nombre del archivo PDF
    timestamp = int(time.time())
    filename = f"Resumen-general-{name}-{timestamp}.pdf"

    # Crear PDF usando WeasyPrint
    pdf = HTML(string=rendered_html).write_pdf()

    # Guardar el PDF en un archivo
    with open(filename, 'wb') as f:
        f.write(pdf)

    # Subir el PDF a S3
    s3.upload_file(filename, s3_bucket, filename, ExtraArgs={'ACL': 'public-read'})
    s3_url = s3.generate_presigned_url(
        ClientMethod="get_object",
        Params={"Bucket": s3_bucket, "Key": filename}
    )

    # Eliminar el archivo local
    os.remove(filename)
    
    return s3_url.split("?")[0]



def buil_pdf_filtered_adq(data, year_adq, method_adq, name="anonimo"):
    # Plantilla HTML
    template_list = Environment().from_string("""
    <html>
    <head>
        <style>
            @page {
                size: A4 landscape; /* Establecer tamaño de página A4 en formato horizontal */
                margin: 6mm; /* Margen opcional */
            }
            @font-face {
                font-family: 'Montserrat';
                src: url('Montserrat-Regular.ttf') format('truetype');
            }
            body {
                font-family: 'Montserrat', sans-serif;
                font-size: 14px;
            }
            table { width: 100%; border-collapse: collapse; }
            th {
                border-bottom: 1px solid black;
                padding: 8px;
                text-align: center;
                vertical-align: center;
                width: 100px;
            }
            td {
                border-bottom: 1px solid black;
                padding: 8px;
                text-align: center;
                vertical-align: center;
                overflow-wrap: break-word;
                text-overflow: elipsis;
                width: 100px;
                height: 90px;
            }
            img { height: 80px; max-width: 80px; /* Limitar tamaño de la imagen */ }
        </style>
    </head>
    <body>
        <sup><i>{{ now }}</i></sup>
        <h1>Resumen de Obras adquiridas en el año {{ year_adq }} en {{ method_adq }}</h1>
        <table>
            <thead>
                <tr>
                    <th>Imagen</th>
                    <th>Artista</th>
                    <th>Medio</th>
                    <th>Nombre</th>
                    <th>Descripción</th>
                    <th>Año</th>
                    <th>Año de adquisición</th>
                    <th>Pais</th>
                    <th>Medidas</th>
                    <th>Ubicación obra</th>
                </tr>
            </thead>
            <tbody>
                {% for obra in obras %}
                    <tr>
                        <td><img src="{{ obra.imagen | escape }}" alt="imagen"></td>
                        <td>{{ obra.artista }}</td>
                        <td>{{ obra.medio }}</td>
                        <td>{{ obra.nombre }}</td>
                        <td>{{ obra.descripcion }}</td>
                        <td style="max-width: 70px;">{{ obra.año }}</td>
                        <td style="max-width: 70px;">{{ obra.año_adquisicion }}</td>
                        <td>{{ obra.pais }}</td>
                        <td style="max-width: 70px;">{{ obra.medidas }}</td>
                        <td>{{ obra.ubicacion_obra }}</td>
                    </tr>
                {% endfor %}
            </tbody>
        </table>
    </body>
    </html>
    """)

    obras = []
    for item in data:
        image_url = item.get('photo_thumb', 'https://qullqua-uploads.s3.amazonaws.com/blanco2.jpeg') or 'https://qullqua-uploads.s3.amazonaws.com/blanco2.jpeg'
        encoded_image_url = quote(image_url, safe='/:')
        description = item.get('description', '')
        description = description[:50] if description else ''
        medidas = item.get('sizes_des', '')
        medidas = medidas[:40] if medidas else ''
        date_adq = item.get('date_adquisition', '')
        year_of_adq = date_adq.split('-')[0] if date_adq else ''

        item2 = {
            "descripcion": description,
            "medidas": medidas,
            "imagen": encoded_image_url,
            "medio": item.get('tecnique_name', ''),
            "artista": item.get('author', ''),
            "nombre": item.get('atname', ''),
            "año": item.get('year_at', ''),
            "año_adquisicion": year_of_adq,
            "pais": item.get('country_name', ''),
            "ubicacion_obra": item.get('artwork_location', '')
        }
        obras.append(item2)

    # Ordenar obras por artista (de A a Z)
    obras.sort(key=lambda x: (x['artista'] or '').lower())

    # Generar HTML con Jinja2
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    display_method = method_adq if method_adq else 'Todas'
    rendered_html = template_list.render(obras=obras, now=now, year_adq=year_adq, method_adq=display_method)

    # Generar nombre del archivo PDF
    timestamp = int(time.time())
    filename = f"Resumen-filtrado-adq-{name}-{timestamp}.pdf"

    # Crear PDF usando WeasyPrint
    pdf = HTML(string=rendered_html).write_pdf()

    # Guardar el PDF en un archivo
    with open(filename, 'wb') as f:
        f.write(pdf)

    # Subir el PDF a S3
    s3.upload_file(filename, s3_bucket, filename, ExtraArgs={'ACL': 'public-read'})
    s3_url = s3.generate_presigned_url(
        ClientMethod="get_object",
        Params={"Bucket": s3_bucket, "Key": filename}
    )

    # Eliminar el archivo local
    os.remove(filename)

    return s3_url.split("?")[0]

def buil_pdf_adq(data, name="anonimo"):
    
    # Plantilla HTML
    template_list = Environment().from_string("""
    <html>
    <head>
        <style>
            @page {
                size: A4 landscape; /* Establecer tamaño de página A4 en formato horizontal */
                margin: 6mm; /* Margen opcional */
            }
            @font-face {
                font-family: 'Montserrat';
                src: url('Montserrat-Regular.ttf') format('truetype'); 
            }
            body {
                font-family: 'Montserrat', sans-serif;
                font-size: 14px; 
            }
            table { width: 100%; border-collapse: collapse; }
            th {
                border-bottom: 1px solid black;
                padding: 8px; 
                text-align: center; 
                vertical-align: center; 
                width: 100px;
            }
            td { 
                border-bottom: 1px solid black;
                padding: 8px; 
                text-align: center; 
                vertical-align: center; 
                overflow-wrap: break-word;
                text-overflow: elipsis;
                width: 100px;
                height: 90px;
            }
            img { height: 80px; max-width: 80px; /* Limitar tamaño de la imagen */ }
        </style>
    </head>
    <body>
        <div>
            <sup><i>{{ now }}</i></sup>
            <h1>Resumen de Obras</h1>
        </div>
        <table>
            <thead>
                <tr>
                    <th>Imagen</th>
                    <th>Artista</th> 
                    <th>Medio</th> 
                    <th>Nombre</th> 
                    <th>Adquisición</th>
                    <th>Fecha de Adquisición</th>
                    <th>País</th> 
                    <th>Medidas</th>
                    <th>Ubicación obra</th>
                </tr>
            </thead>
            <tbody>
                {% for obra in obras %}
                    <tr>
                        <td><img src="{{ obra.imagen | escape }}" alt="imagen"></td>
                        <td>{{ obra.artista }}</td>
                        <td>{{ obra.medio }}</td>
                        <td>{{ obra.nombre }}</td>
                        <td>{{ obra.adquisition }}</td>
                        <td>{{ obra.date_adquisition }}</td>
                        <td>{{ obra.pais }}</td>
                        <td style="max-width: 70px;">{{ obra.medidas }}</td>
                        <td>{{ obra.ubicacion_obra }}</td>
                    </tr>
                {% endfor %}                       
            </tbody>
        </table>
    </body>
    </html>
    """)
    
    # Preparar los datos de las obras
    obras = []
    for item in data:
        image_url = item.get('photo_thumb', 'https://qullqua-uploads.s3.amazonaws.com/blanco2.jpeg') or 'https://qullqua-uploads.s3.amazonaws.com/blanco2.jpeg'
        encoded_image_url = quote(image_url, safe="/:")
        medidas = item.get('sizes_des', '')
        medidas = medidas[:40] if medidas else ''
        item2 = {
            "imagen": encoded_image_url,
            "medidas": medidas,
            "medio": item.get("tecnique_name", ""),
            "nombre": item.get("atname", ""),
            "adquisition": item.get("adquisition", ""),
            "date_adquisition": item.get("date_adquisition", ""),
            "pais": item.get("country_name", ""),
            "artista": item.get("author", ""),
            "ubicacion_obra": item.get("artwork_location", "")
        }
        obras.append(item2)

    obras.sort(key=lambda x: (x['artista'] or '').lower())

    # Renderizar HTML con Jinja2
    now = time.strftime("%Y-%m-%d %H:%M:%S")
    rendered_html = template_list.render(obras=obras, now=now)

    # Generar el nombre del archivo PDF
    timestamp = int(time.time())
    filename = f"Resumen-general-adquisicion-{name}-{timestamp}.pdf"

    # Crear PDF usando WeasyPrint
    pdf = HTML(string=rendered_html).write_pdf(stylesheets=[])

    # Guardar el PDF en un archivo local
    with open(filename, 'wb') as f:
        f.write(pdf)

    # Subir el PDF a S3
    s3.upload_file(filename, s3_bucket, filename, ExtraArgs={"ACL": "public-read"})
    s3_url = s3.generate_presigned_url(
        ClientMethod="get_object",
        Params={"Bucket": s3_bucket, "Key": filename}
    )

    # Eliminar el archivo local
    os.remove(filename)
    
    return s3_url.split("?")[0]


@app.get("/summary/pdf")
async def read_root(token: str = None, sumary : str = '1', limit : int = 10000, year_adq: str = None, method_adq: str = None):
    if token == None:
        result = table.scan(    
        )
        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
            )
            items.extend(result['Items'])
    else:
        result = table.scan(
            FilterExpression=Attr('token').eq(token),
        )

        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
                FilterExpression=Attr('token').eq(token),
            )
            items.extend(result['Items'])
    items = items[:limit]

    if year_adq or method_adq:
        filtered_items = []
        for item in items:
            match_year = True
            match_method = True

            if year_adq:
                date_adq = item.get('date_adquisition')
                if not date_adq or not date_adq.startswith(str(year_adq)):
                    match_year = False

            if method_adq:
                if item.get('adquisition') != method_adq:
                    match_method = False

            if match_year and match_method:
                filtered_items.append(item)
        items = filtered_items

    for item in items:

        tecnique_name = "N/A"
        if 'id_technique' in item:
            if item['id_technique']  is not None and item['id_technique']  != "":
                result_field_tec = table_tecnique.get_item(
                            Key={
                                'id': item['id_technique']
                            }
                        )
                if 'Item' in result_field_tec:
                    tecnique_name = result_field_tec['Item']['tec_name']

        country_name = ""
        if 'country' in item:
            if item['country']  is not None and item['country']  != "":
                result_field_cout = table_country.get_item(
                            Key={
                                'id': item['country']
                            }
                        )
                country_name = result_field_cout['Item']['cname'] if 'Item' in result_field_cout else ""

        item.update(

            {   
                "tecnique_name" : tecnique_name,
                "country_name" : country_name
            }
        )
    
    result_user = table_user.scan(
        FilterExpression=Attr('token').eq(token) 
    )
    if token is not None:
        user = "NA"
        if 'Items' in result_user and result_user['Items']:
            user = result_user['Items'][0].get('first_name','').replace(" ","-")
    else:
        user = "admin"
        
    if sumary == '3':
        url = buil_pdf_filtered_adq(items, year_adq, method_adq, user)
    elif sumary == '2':
        url = buil_pdf_adq(items, user)
    else:  
        url = buil_pdf(items, user)
    return {"url": url}


@app.get("/inventory/location")
async def read_root(token: str = None):
    
    if token == None:
        result = table.scan(    
        )
        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
            )
            items.extend(result['Items'])
    else:
        result = table.scan(
            FilterExpression=Attr('token').eq(token),
        )

        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
                FilterExpression=Attr('token').eq(token) 
            )
            items.extend(result['Items'])

    autores_por_ubicacion = {}

    
    for obra in items:
        autor = obra.get("author", "Anonimo")
        ubicacion = obra.get("artwork_location", "N/A")

        if ubicacion not in autores_por_ubicacion:
            autores_por_ubicacion[ubicacion] = {}

        if autor not in autores_por_ubicacion[ubicacion]:
            autores_por_ubicacion[ubicacion][autor] = 0

        autores_por_ubicacion[ubicacion][autor] += 1

    resultado_final = {}

    for ubicacion, autores in autores_por_ubicacion.items():
        resultado_final[ubicacion] = [{"name": autor, "count": count} for autor, count in autores.items()]

    result_user = table_user.scan(
        FilterExpression=Attr('token').eq(token) 
                )
    
    user = result_user['Items'][0].get('first_name','').strip()
    url = buil_pdf_inventory(resultado_final,'artistas', user)

    return {"url": url}



@app.get("/inventory/medio")
async def read_root(token: str = None):
    
    if token == None:
        result = table.scan(    
        )
        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
            )
            items.extend(result['Items'])
    else:
        result = table.scan(
            FilterExpression=Attr('token').eq(token),
        )

        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
                FilterExpression=Attr('token').eq(token) 
            )
            items.extend(result['Items'])

  

    for item in items:
            tecnique_name = ""
            if 'id_technique' in item:
                if item['id_technique']  is not None and item['id_technique']  != "":
                    result_field_cout = table_tecnique.get_item(
                                Key={
                                    'id': item['id_technique']
                                }
                            )
                    if 'Item' in result_field_cout:
                        tecnique_name = result_field_cout['Item']['tec_name'] if len(result_field_cout['Item']) >0 else 'N/A'
                    else:
                         tecnique_name = "N/A"
                else:
                    tecnique_name = "N/A"
           
            item.update({"tecnique" : tecnique_name})

 
    autores_por_ubicacion = {}

    for obra in items:
        tecnica = obra.get("tecnique", "N/A")
        ubicacion = obra.get("artwork_location", "N/A")

        if ubicacion not in autores_por_ubicacion:
            autores_por_ubicacion[ubicacion] = {}

        if tecnica not in autores_por_ubicacion[ubicacion]:
            autores_por_ubicacion[ubicacion][tecnica] = 0

        autores_por_ubicacion[ubicacion][tecnica] += 1

    resultado_final = {}

    for ubicacion, tecnicas in autores_por_ubicacion.items():
        resultado_final[ubicacion] = [{"name": tecnica, "count": count} for tecnica, count in tecnicas.items()]

    result_user = table_user.scan(
        FilterExpression=Attr('token').eq(token) 
                )
    
    user = result_user['Items'][0].get('first_name','').strip()

    url = buil_pdf_inventory(resultado_final,'medio',user)

    return {"url": url}

@app.get("/excel/simple")
async def read_root(token: str = None):
    
    if token == None:
        result = table.scan(    
        )
        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
            )
            items.extend(result['Items'])
    else:
        result = table.scan(
            FilterExpression=Attr('token').eq(token),
        )

        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
                FilterExpression=Attr('token').eq(token) 
            )
            items.extend(result['Items'])
    for item in items:

            tecnique_name = "N/A"
            if 'id_technique' in item:
                if item['id_technique']  is not None and item['id_technique']  != "":
                    result_field_tec = table_tecnique.get_item(
                                Key={
                                    'id': item['id_technique']
                                }
                            )
                    tecnique_name = result_field_tec['Item']['tec_name']
                    
                else:
                    tecnique_name = "N/A"

            category_name = "N/A"
            if 'id_category' in item:
                if item['id_category']  is not None and item['id_category']  != "":
                    result_field_cat = table_category.get_item(
                                Key={
                                    'id': item['id_category']
                                }
                            )
                    category_name = result_field_cat['Item']['cat_name']
                    
                else:
                    category_name = "N/A"

            item.update(

                {   
                    "tecnique_name" : tecnique_name,      
                    "category_name" : category_name
         
                }
            )
    result_user = table_user.scan(
        FilterExpression=Attr('token').eq(token) 
                )
    
    user = result_user['Items'][0].get('first_name','').replace(" ","-")

    s3_url = build_simple_xls(items,"1",user)

    return {"url": s3_url}



@app.get("/excel/todo")
async def read_root(token: str = None):
    
    if token == None:
        result = table.scan(    
        )
        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
            )
            items.extend(result['Items'])
    else:
        result = table.scan(
            FilterExpression=Attr('token').eq(token),
        )

        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
                FilterExpression=Attr('token').eq(token) 
            )
            items.extend(result['Items'])
    for item in items:
            country_name = ""
            if 'country' in item:
                if item['country']  is not None and item['country']  != "":
                    result_field_cout = table_country.get_item(
                                Key={
                                    'id': item['country']
                                }
                            )
                    country_name = result_field_cout['Item']['cname'] if 'Item' in result_field_cout else ""
                    
                else:
                    country_name = ""
            else:
                country_name = ""


            tecnique_name = "N/A"
            if 'id_technique' in item:
                if item['id_technique']  is not None and item['id_technique']  != "":
                    result_field_tec = table_tecnique.get_item(
                                Key={
                                    'id': item['id_technique']
                                }
                            )
                    tecnique_name = result_field_tec['Item']['tec_name']
                    
                else:
                    tecnique_name = "N/A"

            category_name = "N/A"
            if 'id_category' in item:
                if item['id_category']  is not None and item['id_category']  != "":
                    result_field_cat = table_category.get_item(
                                Key={
                                    'id': item['id_category']
                                }
                            )
                    category_name = result_field_cat['Item']['cat_name']
                    
                else:
                    category_name = "N/A"

            collection_name = "N/A"
            if 'id_colection' in item:
                if item['id_colection']  is not None and item['id_colection']  != "":
                    result_field_co = table_coll.get_item(
                                Key={
                                    'id': item['id_colection']
                                }
                            )
                    if 'Item' in result_field_co:
                        collection_name = result_field_co['Item']['collname']
                    else:
                         collection_name = "N/A"
                    
                else:
                    collection_name = "N/A"

            if 'photo' in item:
                if isinstance(item['photo'], list):
                    if len(item['photo'])> 0:
                        photo = item['photo'][0]
                    else:
                        photo = ""
                else:
                    photo = item['photo']
            item.update(

                {   
                    "tecnique_name" : tecnique_name,      
                    "category_name" : category_name,
                    "country_name": country_name,
                    "collection_name": collection_name,
                    "photo" : photo
         
                }
            )
    result_user = table_user.scan(
        FilterExpression=Attr('token').eq(token) 
                )
    
    user = result_user['Items'][0].get('first_name','').replace(" ","-")

    s3_url = build_simple_xls(items,"2",user)

    return {"url": s3_url}



@app.get("/artwoks/order")
async def read_root(token: str = None, field: str  = None, order: str= "ASC" ):
    
    if token == None:
        result = table.scan(    
        )
        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
            )
            items.extend(result['Items'])
    else:
        result = table.scan(
            FilterExpression=Attr('token').eq(token),
        )

        items = result['Items']
        while 'LastEvaluatedKey' in result:
            result = table.scan(
                ExclusiveStartKey=result['LastEvaluatedKey'],
                FilterExpression=Attr('token').eq(token) 
            )
            items.extend(result['Items'])
    for item in items:
            country_name = ""
            if 'country' in item:
                if item['country']  is not None and item['country']  != "":
                    result_field_cout = table_country.get_item(
                                Key={
                                    'id': item['country']
                                }
                            )
                    country_name = result_field_cout['Item']['cname'] if 'Item' in result_field_cout else ""
                    
                else:
                    country_name = ""
            else:
                country_name = ""


            tecnique_name = "N/A"
            if 'id_technique' in item:
                if item['id_technique']  is not None and item['id_technique']  != "":
                    result_field_tec = table_tecnique.get_item(
                                Key={
                                    'id': item['id_technique']
                                }
                            )
                    tecnique_name = result_field_tec['Item']['tec_name']
                    
                else:
                    tecnique_name = "N/A"

            category_name = "N/A"
            if 'id_category' in item:
                if item['id_category']  is not None and item['id_category']  != "":
                    result_field_cat = table_category.get_item(
                                Key={
                                    'id': item['id_category']
                                }
                            )
                    category_name = result_field_cat['Item']['cat_name']
                    
                else:
                    category_name = "N/A"

            collection_name = "N/A"
            if 'id_colection' in item:
                if item['id_colection']  is not None and item['id_colection']  != "":
                    result_field_co = table_coll.get_item(
                                Key={
                                    'id': item['id_colection']
                                }
                            )
                    if 'Item' in result_field_co:
                        collection_name = result_field_co['Item']['collname']
                    else:
                         collection_name = "N/A"
                    
                else:
                    collection_name = "N/A"

            if 'photo' in item:
                if isinstance(item['photo'], list):
                    if len(item['photo'])> 0:
                        photo = item['photo'][0]
                    else:
                        photo = ""
                else:
                    photo = item['photo']
            item.update(

                {   
                    "tecnique_name" : tecnique_name,      
                    "category_name" : category_name,
                    "country_name": country_name,
                    "collection_name": collection_name,
                    "photo" : photo
         
                }
            )
   
    if order == "ASC":
        control = False
    else: 
        control = True

    items3 = sorted(items,key=lambda x: (x[field] is None, x[field] == "", x[field]),reverse=control)

    data = {   
            "status": "ok",
            "total_artworks":len(items),
            "data": items3,
            "total_wishlist": "0",
            "pay_name": "Collector Pro",
            "pay_code": "3",
            "pay_max_items": "500",
            "pay_date_init": "2020-07-21 11:19:28",
            "pay_date_end": "2024-02-09 07:52:13"
        }


    return {"data":data}
    

@app.get("/health")
async def read_root():
    return {"message":"estoy vivo"}

if __name__ == '__main__':
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
